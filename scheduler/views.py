# scheduler/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.db import transaction
from scheduling.models import Schedule, Semester, Section, Room, Curriculum, GenEdSchedule, InstructorSchedulingConfiguration
from core.models import Instructor, UserLogin, User
from aimatching.models import InstructorSubjectMatch
from django.contrib.auth.decorators import login_required
from collections import defaultdict
from django.http import JsonResponse
from scheduler.tasks import run_scheduler_task
from scheduler.models import SchedulerProgress
from celery import current_app
import uuid
import os, signal
from django.contrib import messages
from datetime import datetime, time, timedelta
from django.utils import timezone
from authapi.views import has_role  
import re
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from django.db.models import Sum
from django.template.loader import render_to_string
import math
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.conf import settings
from openpyxl.cell.cell import MergedCell
from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import json
from django.db.models import Q
from django.core.cache import cache
from scheduling.models import Semester



@login_required
@has_role('deptHead')
def finalizeSchedule(request, semester_id): 
    if request.method == 'POST':
        active_schedules = Schedule.objects.filter(semester__semesterId=semester_id, status='active')
        
        if active_schedules.exists():
            try:
                with transaction.atomic():
                    Schedule.objects.filter(semester__semesterId=semester_id, status='finalized').update(status='archived')
                    
                    active_schedules.update(status='finalized')
                    
                    messages.success(request, "Schedule successfully FINALIZED and locked.")
            except Exception as e:
                messages.error(request, f"Failed to finalize schedule: {e}")
        else:
            messages.warning(request, "No active schedule found to finalize.")
            
        return redirect(reverse('scheduleOutput') + f'?semester={semester_id}&batch_key=finalized')
    return redirect('scheduleOutput')

@login_required
@has_role('deptHead')
def revertFinalizedSchedule(request, semester_id):
    if request.method == 'POST':
        try:
            semester = Semester.objects.get(semesterId=semester_id)
        except Semester.DoesNotExist:
            messages.error(request, "Semester not found.")
            return redirect(reverse('scheduleOutput'))

        try:
            with transaction.atomic():
                finalized_schedules = Schedule.objects.filter(
                    semester=semester, 
                    status='finalized'
                )

                if finalized_schedules.exists():
                    finalized_schedules.update(status='active')
                    messages.success(request, f"Schedule for {semester.name} has been **UNLOCKED**.")
                else:
                    messages.warning(request, "No finalized schedule found to unlock.")
        except Exception as e:
            messages.error(request, f"An error occurred while unlocking the schedule: {e}")
        return redirect(reverse('scheduleOutput') + f'?semester={semester_id}&batch_key=active')
    return redirect(reverse('scheduleOutput'))


@login_required
@has_role('deptHead')
def revertSchedule(request):
    if request.method == 'POST':
        semester_id = request.POST.get('semester_id')
        batch_key = request.POST.get('batch_key') 

        if not semester_id or not batch_key or batch_key in ['active', 'finalized']:
            messages.error(request, "Invalid request for schedule reversion.")
            return redirect(reverse('scheduleOutput'))
        
        finalized_run_exists = Schedule.objects.filter(
            semester__semesterId=semester_id, 
            status='finalized'
        ).exists()
        
        if finalized_run_exists:
            messages.error(request, "A schedule is currently **FINALIZED**. Please unlock it first.")
            return redirect(reverse('scheduleOutput') + f'?semester={semester_id}&batch_key=finalized')
        
        try:
            dt_obj = datetime.fromisoformat(batch_key)
            
            if dt_obj.tzinfo is None:
                dt_obj = timezone.make_aware(dt_obj)
            
            schedules_to_promote = Schedule.objects.filter(
                semester__semesterId=semester_id,
                status='archived',
                createdAt__gte=dt_obj,
                createdAt__lt=dt_obj + timedelta(seconds=1)
            ).select_related('semester')
            
            if not schedules_to_promote.exists():
                messages.error(request, "The specified archived schedule batch was not found.")
                return redirect(reverse('scheduleOutput') + f'?semester={semester_id}&batch_key=active')
            
            semester = schedules_to_promote.first().semester

            with transaction.atomic():
                Schedule.objects.filter(semester=semester, status='active').update(status='archived')
                schedules_to_promote.update(status='active')
                messages.success(request, f"Successfully reverted to schedule batch created at {batch_key}.")

        except ValueError:
            messages.error(request, "Invalid batch key format.")
        except Exception as e:
            messages.error(request, f"Error during reversion: {e}")
        
        return redirect(reverse('scheduleOutput') + f'?semester={semester_id}&batch_key=active')
    
    return redirect(reverse('scheduleOutput'))


@login_required
@has_role('deptHead')
def scheduleOutput(request):
    semester_id = request.GET.get("semester")
    if not semester_id:
        latest_semester = Semester.objects.order_by("-createdAt").first()
        if latest_semester:
            semester_id = str(latest_semester.semesterId)
        else:
            return render(request, "scheduler/scheduleOutput.html", {"error": "No semesters found."})

    finalized_exists = Schedule.objects.filter(
        semester__semesterId=semester_id, 
        status='finalized'
    ).exists()

    raw_batch_key = request.GET.get("batch_key")

    if finalized_exists and (raw_batch_key is None or raw_batch_key == 'active'):
        return redirect(reverse('scheduleOutput') + f'?semester={semester_id}&batch_key=finalized')

    batch_key = raw_batch_key if raw_batch_key else 'active'
    
    raw_archived_datetimes = Schedule.objects.filter(
        semester__semesterId=semester_id, 
        status='archived'
    ).order_by('-createdAt').values_list('createdAt', flat=True).distinct()
    
    unique_batch_times = {}
    for dt in raw_archived_datetimes:
        truncated_dt_key = dt.replace(microsecond=0)
        if truncated_dt_key not in unique_batch_times:
            unique_batch_times[truncated_dt_key] = truncated_dt_key
            
    archived_batch_times = sorted(unique_batch_times.keys(), reverse=True)

    archived_batches = [
        {'key': batch_time.isoformat(), 'label': f"Archived Run: {batch_time.strftime('%b %d, %Y %I:%M %p')}"} 
        for batch_time in archived_batch_times
    ]

    schedules = Schedule.objects.none() 
    current_status = 'N/A'
    
    if batch_key == 'finalized' and finalized_exists:
        schedules = Schedule.objects.filter(semester__semesterId=semester_id, status='finalized')
        current_status = 'Finalized Schedule (Locked)'
    
    elif batch_key == 'active':
        schedules = Schedule.objects.filter(semester__semesterId=semester_id, status='active')
        current_status = 'Active Draft'

        if not schedules.exists() and archived_batches:
             try:
                 latest_batch_time_iso = archived_batches[0]['key']
                 latest_batch_time = datetime.fromisoformat(latest_batch_time_iso)
                 if latest_batch_time.tzinfo is None:
                     latest_batch_time = timezone.make_aware(latest_batch_time)

                 schedules = Schedule.objects.filter(
                     semester__semesterId=semester_id, 
                     createdAt__gte=latest_batch_time,
                     createdAt__lt=latest_batch_time + timedelta(seconds=1),
                     status='archived'
                 )
                 current_status = archived_batches[0]['label']
                 batch_key = latest_batch_time_iso
                 messages.info(request, "Active draft not found. Displaying the latest archived run instead.")
             except Exception:
                 schedules = Schedule.objects.none()
                 current_status = 'No Schedules Found'
                 batch_key = 'none'

    elif batch_key:
        try:
            selected_timestamp = datetime.fromisoformat(batch_key)
            if selected_timestamp.tzinfo is None:
                selected_timestamp = timezone.make_aware(selected_timestamp)
            selected_timestamp = selected_timestamp.replace(microsecond=0)
            
            schedules = Schedule.objects.filter(
                semester__semesterId=semester_id, 
                createdAt__gte=selected_timestamp,
                createdAt__lt=selected_timestamp + timedelta(seconds=1),
                status='archived'
            )
            label = next((b['label'] for b in archived_batches if b['key'] == batch_key), f"Archived Run ({selected_timestamp.strftime('%b %d, %I:%M %p')})")
            current_status = label

        except ValueError:
            messages.error(request, "Invalid schedule batch key provided. Falling back to active draft.")
            return redirect(reverse('scheduleOutput') + f'?semester={semester_id}&batch_key=active')
    
    schedules = schedules.select_related(
        'instructor', 'room'
    ).prefetch_related(
        'instructor__userlogin_set__user'
    ).order_by('dayOfWeek', 'startTime')
    
    DAYS_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    
    room_usage = defaultdict(lambda: {
        "room_code": None, "total_minutes": 0, "usage_hours": 0.0,
        "daily_spread": defaultdict(int)
    })
    
    instructor_load = defaultdict(lambda: {
        "id": None, "name": None, "normal_min": 0, "overload_min": 0, "total_hours": 0.0,
        "daily_spread": defaultdict(int)
    })
    
    for schedule in schedules:
        duration_minutes = schedule.duration_minutes 
        
        room_name = schedule.room.roomCode if schedule.room else "TBA"
        room_usage[room_name]["room_code"] = room_name
        room_usage[room_name]["total_minutes"] += duration_minutes
        room_usage[room_name]["daily_spread"][schedule.dayOfWeek] += duration_minutes
        
        if schedule.instructor:
            instr_id = schedule.instructor.instructorId 
            instr_name = schedule.instructor.full_name
        else:
            instr_id = 'TBA_UNASSIGNED' 
            instr_name = 'Unassigned Sections'

        instructor_load[instr_id]["id"] = instr_id
        instructor_load[instr_id]["name"] = instr_name
        
        if schedule.isOvertime: 
            instructor_load[instr_id]["overload_min"] += duration_minutes
        else:
            instructor_load[instr_id]["normal_min"] += duration_minutes
            
        instructor_load[instr_id]["daily_spread"][schedule.dayOfWeek] += duration_minutes

    formatted_room_data = [] 
    for room_code, data in room_usage.items():
        data["usage_hours"] = round(data["total_minutes"] / 60.0, 2)
        daily_list = []
        for day in DAYS_ORDER:
            daily_list.append({"day": day, "hours": round(data["daily_spread"].get(day, 0) / 60.0, 2)})
        data["daily_spread_list"] = daily_list
        formatted_room_data.append(data)
    ordered_room_usage = sorted(formatted_room_data, key=lambda item: item["usage_hours"], reverse=True)

    formatted_instructor_data = []
    for instr_id, data in instructor_load.items(): 
        if instr_id == 'TBA_UNASSIGNED': 
            continue 
        data["total_hours"] = round((data["normal_min"] + data["overload_min"]) / 60.0, 2)
        data["normal_hours"] = round(data["normal_min"] / 60.0, 2)
        data["overload_hours"] = round(data["overload_min"] / 60.0, 2)
        daily_list = []
        for day in DAYS_ORDER:
            daily_list.append({"day": day, "hours": round(data["daily_spread"].get(day, 0) / 60.0, 2)})
        data["daily_spread_list"] = daily_list
        formatted_instructor_data.append(data)
    ordered_instructor_load = sorted(formatted_instructor_data, key=lambda item: item["total_hours"], reverse=True)

    semesters = Semester.objects.all().order_by("-createdAt")
    try:
        current_semester = Semester.objects.get(semesterId=semester_id)
    except Semester.DoesNotExist:
        current_semester = None

    context = {
        "current_semester_id": semester_id,
        "current_status": current_status,
        "batch_key": batch_key,
        "finalized_exists": finalized_exists,
        "schedules": schedules,
        "semesters": semesters,
        "archived_batches": archived_batches,
        "current_semester": current_semester,
        "room_usage_data": ordered_room_usage,
        "instructor_load_data": ordered_instructor_load,
        "days_order": DAYS_ORDER, 
        
        "can_finalize": (batch_key == 'active' and schedules.exists() and not finalized_exists),
        "is_finalized": current_status.startswith('Finalized'),
        "can_revert": (batch_key not in ['active', 'finalized', 'none', 'N/A']) and schedules.exists(), 
    }
    return render(request, "scheduler/scheduleOutput.html", context)


@login_required
@has_role('instructor')
def instructorScheduleView(request):
    user = request.user
    
    login_entry = UserLogin.objects.filter(user=user).select_related("instructor").first()
    if not login_entry or not login_entry.instructor:
        return render(request, "scheduler/instructorSchedule.html", {
            "error": "No instructor profile found for this account."
        })

    instructor = login_entry.instructor

    all_semesters = Semester.objects.order_by('-academicYear', '-term')
    semester_id = request.GET.get("semester")
    
    current_semester = None
    if semester_id:
        current_semester = all_semesters.filter(semesterId=semester_id).first()
    
    if not current_semester:
        current_semester = all_semesters.first()
        if current_semester:
            semester_id = str(current_semester.semesterId)

    if not current_semester:
         return render(request, "scheduler/instructorSchedule.html", {
            "error": "No semesters found in the system.",
            "instructor": instructor
        })

    schedules = Schedule.objects.filter(
        instructor=instructor,
        semester=current_semester,
        status='finalized' 
    ).select_related("subject", "section", "room")

    schedule_list = list(schedules)
    
    start_hour_floor = 7 
    end_hour_ceiling = 19

    if schedule_list:
        earliest_class_hour = min(s.startTime.hour for s in schedule_list)
        latest_class_hour = max(s.endTime.hour for s in schedule_list)
        
        if 0 < earliest_class_hour < start_hour_floor:
            start_hour_floor = earliest_class_hour
        
        if latest_class_hour > end_hour_ceiling:
            end_hour_ceiling = latest_class_hour

    min_time = time(max(0, start_hour_floor - 1), 0)
    max_time = time(min(23, end_hour_ceiling + 1), 0)

    start_dt = datetime.combine(datetime.today(), min_time)
    if start_dt.minute >= 30:
        start_dt = start_dt.replace(minute=30, second=0, microsecond=0)
    else:
        start_dt = start_dt.replace(minute=0, second=0, microsecond=0)

    end_dt = datetime.combine(datetime.today(), max_time)
    if end_dt.minute > 30:
        end_dt = end_dt + timedelta(hours=1).replace(minute=0, second=0, microsecond=0)
    elif end_dt.minute > 0:
        end_dt = end_dt.replace(minute=30, second=0, microsecond=0)
    
    time_slots_display = []
    current_dt = start_dt
    while current_dt < end_dt:
        time_slots_display.append(current_dt.strftime('%I:%M %p'))
        current_dt += timedelta(minutes=30)
    
    DAYS_OF_WEEK = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    raw_data = defaultdict(lambda: defaultdict(list))

    color_palette = [
        'bg-blue-600', 'bg-red-600', 'bg-emerald-600', 'bg-purple-600', 
        'bg-orange-600', 'bg-teal-600', 'bg-pink-600', 'bg-indigo-600', 
        'bg-cyan-600', 'bg-rose-600', 'bg-amber-600', 'bg-lime-600'
    ]
    section_color_map = {}
    color_index = 0

    for s in schedule_list:
        if s.dayOfWeek not in DAYS_OF_WEEK:
            continue
        
        sched_start_dt = datetime.combine(datetime.today(), s.startTime)
        sched_end_dt = datetime.combine(datetime.today(), s.endTime)
        if sched_end_dt < sched_start_dt:
             sched_end_dt += timedelta(days=1)
        
        subject_code = s.subject.code
        year_level = ""
        match_year = re.search(r'\d', subject_code)
        if match_year:
            year_level = match_year.group()
        
        raw_section = s.section.sectionCode
        if '-' in raw_section:
            section_letter = raw_section.split('-')[-1].strip()
        else:
            section_letter = raw_section.strip()
            
        section_str = f"{year_level}{section_letter}" 


        raw_type = str(getattr(s, 'scheduleType', 'lecture')).lower() # Fixed attr name based on your model
        type_label = "Lec"
        if "lab" in raw_type:
            type_label = "Lab"


        unique_key = f"{subject_code}_{section_str}"
        if unique_key not in section_color_map:
            section_color_map[unique_key] = color_palette[color_index % len(color_palette)]
            color_index += 1
        
        assigned_color_class = section_color_map[unique_key]
        assigned_hover_class = assigned_color_class.replace('600', '700')

        subject_str = s.subject.code
        room_str = s.room.roomCode if s.room else "TBA"
        
        current_slot_dt = start_dt
        slot_index = 0
        
        while current_slot_dt < end_dt:
            slot_end_dt = current_slot_dt + timedelta(minutes=30)
            slot_time_str = time_slots_display[slot_index]

            if sched_start_dt < slot_end_dt and sched_end_dt > current_slot_dt:
                rowspan = int(s.duration_minutes / 30)
                
                raw_data[s.dayOfWeek][slot_time_str].append({
                    'subject': subject_str,
                    'section': section_str,
                    'room': room_str,
                    'type_label': type_label,
                    'start_time': s.startTime.strftime('%I:%M %p'),
                    'end_time': s.endTime.strftime('%I:%M %p'),
                    'rowspan': rowspan,
                    'height_px': rowspan * 40,
                    'is_start_slot': (sched_start_dt >= current_slot_dt and sched_start_dt < slot_end_dt) or (current_slot_dt == start_dt and sched_start_dt < start_dt),
                    
                    'color_class': assigned_color_class,
                    'hover_class': assigned_hover_class
                })
            
            current_slot_dt = slot_end_dt
            slot_index += 1

    matrix = []
    for time_slot in time_slots_display:
        row_data = {'time': time_slot, 'days': []}
        for day in DAYS_OF_WEEK:
            schedules_in_cell = raw_data[day][time_slot]
            row_data['days'].append(schedules_in_cell)
        matrix.append(row_data)

    context = {
        "instructor": instructor,
        "semesters": all_semesters,
        "current_semester": current_semester,
        "matrix": matrix,
        "days_of_week": DAYS_OF_WEEK,
        "has_schedules": len(schedule_list) > 0,
        "selected_semester_id": semester_id # Pass back to keep dropdown selected
    }
    
    return render(request, "scheduler/instructorSchedule.html", context)


# Scheduler Dashboard View
@login_required
@has_role('deptHead')
def scheduler_dashboard(request):
    batch_id = str(uuid.uuid4())
    return render(request, "scheduler/schedulerDashboard.html", {"batch_id": batch_id})

@login_required
@has_role('deptHead')
def start_scheduler(request):
    batch_id = request.GET.get("batch_id")
    
    semester = Semester.objects.filter(isActive=True).first()
    if not semester:
        return JsonResponse({"status": "error", "message": "No active semester found."}, status=400)

    lock_id = f"scheduler_lock_{semester.semesterId}"
    if cache.get(lock_id):
        # 🚨 BLOCK IT HERE
        return JsonResponse({
            "status": "error", 
            "message": "⚠️ The Scheduler is already running in another tab or device! Please wait for it to finish."
        }, status=423) # 423 = Locked

    # 3. Proceed as normal if no lock
    progress, _ = SchedulerProgress.objects.get_or_create(batch_id=batch_id)
    
    cache.set(lock_id, "starting", timeout=30) 

    task = run_scheduler_task.delay(batch_id=batch_id)
    progress.task_id = task.id
    progress.status = "running"
    progress.save()
    
    return JsonResponse({"message": "Scheduling started.", "batch_id": batch_id, "task_id": task.id})


@login_required
@has_role('deptHead')
def stop_scheduler(request):
    batch_id = request.GET.get("batch_id")
    try:
        progress = SchedulerProgress.objects.get(batch_id=batch_id)

        if progress.process_pid:
            try:
                os.kill(progress.process_pid, signal.SIGTERM)
            except ProcessLookupError:
                pass  

        if progress.task_id:
            current_app.control.revoke(progress.task_id, terminate=True, signal="SIGTERM")

        progress.status = "stopped"
        progress.message = "Scheduler manually stopped."
        progress.save()

        return JsonResponse({"status": "stopped", "message": "Scheduler has been stopped."})
    except SchedulerProgress.DoesNotExist:
        return JsonResponse({"status": "error", "message": "No running scheduler found."})
    

@login_required
@has_role('deptHead')
def scheduler_status(request):
    batch_id = request.GET.get("batch_id")
    if not batch_id:
        return JsonResponse({"error": "Missing batch_id"}, status=400)

    try:
        progress = SchedulerProgress.objects.get(batch_id=batch_id)
    except SchedulerProgress.DoesNotExist:
        return JsonResponse({"status": "idle", "message": "No progress found."})

    return JsonResponse({
        "status": progress.status,
        "message": progress.message,
        "progress": progress.progress,
        "logs": progress.logs[-20:],
    })


def format_number(num):
    try:
        f_num = float(num)
        if f_num.is_integer(): return int(f_num)
        return f_num
    except (ValueError, TypeError): return 0

def get_semester_text(term):
    mapping = {'1st': 'FIRST', '2nd': 'SECOND', 'Midyear': 'MIDYEAR'}
    return mapping.get(term, str(term).upper())

def get_short_day(day_full):
    mapping = {'Monday': 'M', 'Tuesday': 'T', 'Wednesday': 'W', 'Thursday': 'TH', 'Friday': 'F', 'Saturday': 'SAT', 'Sunday': 'SUN'}
    return mapping.get(day_full, day_full)

def get_instructor_involvement(instructor):
    context = {
        'inv_admin': 0,
        'inv_research': 0,
        'inv_extension': 0,
        'inv_consultation': 0,
        'inv_others': 0,
    }

    if instructor.designation:
        d = instructor.designation
        context['inv_admin'] = d.adminSupervisionHours
        context['inv_research'] = d.researchHours
        context['inv_extension'] = d.extensionHours
        context['inv_consultation'] = d.instructionHours + d.consultationHours
        context['inv_others'] = d.productionHours
    elif instructor.rank:
        r = instructor.rank
        context['inv_admin'] = 0 
        context['inv_research'] = r.researchHours
        context['inv_extension'] = r.extensionHours
        context['inv_consultation'] = r.instructionHours + r.consultationHours + r.classAdviserHours
        context['inv_others'] = r.productionHours

    return context

def format_section_time(section):
    schedules = list(section.schedule_set.all())
    schedules.sort(key=lambda x: x.startTime)

    times = []
    use_compact = len(schedules) > 1 

    for s in schedules:
        if use_compact:
            def fmt(t):
                h = t.hour % 12 or 12
                return f"{h}:{t.minute:02d}" if t.minute > 0 else f"{h}"
            
            t_str = f"{fmt(s.start_time)}-{fmt(s.end_time)}"
        else:
            def fmt_std(t):
                return t.strftime("%I:%M")
            t_str = f"{fmt_std(s.start_time)}-{fmt_std(s.end_time)}"
            
        times.append(t_str)

    return " / ".join(times) if times else "TBA"


def process_schedule_group(s_list):
    if not s_list: return None
    
    first = s_list[0]
    subject = first.subject
    
    lec_minutes = getattr(subject, 'durationMinutes', 0)
    lab_minutes = getattr(subject, 'labDurationMinutes', 0) if getattr(subject, 'hasLab', False) else 0
    
    total_lec_hours = lec_minutes / 60
    total_lab_hours = lab_minutes / 60

    sorted_list = sorted(s_list, key=lambda x: (x.dayOfWeek, x.startTime))
    times, days = [], []
    
    use_compact = len(sorted_list) > 1

    for s in sorted_list:
        if use_compact:
            def fmt(t):
                h = t.hour % 12 or 12
                return f"{h}:{t.minute:02d}" if t.minute > 0 else f"{h}"
            t_str = f"{fmt(s.startTime)}-{fmt(s.endTime)}"
        else:
            t_str = f"{s.startTime.strftime('%I:%M')}-{s.endTime.strftime('%I:%M')}"

        times.append(t_str)
        days.append(get_short_day(s.dayOfWeek))

    time_str = " / ".join(times) 
    day_str = " / ".join(days)

    sec_obj = getattr(first, 'section', None)
    year_val = getattr(subject, 'yearLevel', '')
    if year_val is None: year_val = ""

    raw_sec = getattr(sec_obj, 'sectionCode', '') if sec_obj else ""
    if raw_sec:
        sec_letter = raw_sec.split('-')[-1].strip() if '-' in raw_sec else raw_sec.strip()
    else:
        sec_letter = "?" 

    formatted_section = f"BSIT {year_val}{sec_letter}"
    
    student_count = sec_obj.numberOfStudents if (sec_obj and hasattr(sec_obj, 'numberOfStudents')) else 0

    return {
        'code': subject.code,
        'description': subject.name,
        'units': format_number(subject.units),
        'time': time_str,
        'days': day_str,
        'lec': format_number(total_lec_hours) if total_lec_hours > 0 else "", 
        'lab': format_number(total_lab_hours) if total_lab_hours > 0 else 0,
        'students': student_count,
        'room': first.room.roomCode if (hasattr(first, 'room') and first.room) else "TBA",
        'section': formatted_section
    }


def safe_write(ws, row, col, value, align_center=False):
    cell = ws.cell(row=row, column=col)
    target_cell = cell

    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                target_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    
    target_cell.value = value
    
    if align_center:
        target_cell.alignment = Alignment(horizontal='center', vertical='center')

def copy_row_style(ws, source_row_idx, target_row_idx):
    source_row = ws[source_row_idx]
    target_row = ws[target_row_idx]

    for src_cell, tgt_cell in zip(source_row, target_row):
        if src_cell.has_style:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = copy(src_cell.number_format)
            tgt_cell.protection = copy(src_cell.protection)
            tgt_cell.alignment = copy(src_cell.alignment)

def write_section_totals(ws, total_row_idx, data_rows, col_map):
    if not data_rows or not col_map:
        return

    t_units = 0
    t_lec = 0
    t_lab = 0
    t_students = 0

    for row in data_rows:
        def val(k):
            v = row.get(k)
            try:
                if v is None or v == "": return 0
                return float(v)
            except (ValueError, TypeError):
                return 0
        
        t_units += val('units')
        t_lec += val('lec')
        t_lab += val('lab')
        t_students += val('students')

    if 'units' in col_map:
        safe_write(ws, total_row_idx, col_map['units'], t_units, align_center=True)

    if 'lec' in col_map:
        safe_write(ws, total_row_idx, col_map['lec'], t_lec, align_center=True)
        
    if 'lab' in col_map:
        safe_write(ws, total_row_idx, col_map['lab'], t_lab, align_center=True)

    if 'students' in col_map:
        safe_write(ws, total_row_idx, col_map['students'], t_students, align_center=True)

    if 'lec' in col_map:
        grand_total = t_lec + t_lab
        safe_write(ws, total_row_idx + 1, col_map['lec'], grand_total, align_center=True)


def fill_list_section(ws, data_rows, start_search_row=1):
    template_row_idx = None
    col_map = {} 

    for row in ws.iter_rows(min_row=start_search_row):
        possible_map = {}
        found_code = False
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip()
                possible_map[val.lower()] = cell.column
                if val == 'code': found_code = True
        
        if found_code:
            template_row_idx = row[0].row
            col_map = possible_map
            break
    
    if not template_row_idx:
        return start_search_row, {}

    if not data_rows:
        for r_idx in range(template_row_idx + 1, template_row_idx + 20):
            cell_val = ws.cell(row=r_idx, column=2).value
            if cell_val and "TOTAL" in str(cell_val).upper():
                return r_idx, col_map
        return template_row_idx, col_map

    num_to_insert = len(data_rows) - 1
    
    if num_to_insert > 0:
        merges_to_shift = []
        template_row_merges = []
        all_merges = list(ws.merged_cells.ranges)
        
        for merged_range in all_merges:
            if merged_range.min_row == template_row_idx and merged_range.max_row == template_row_idx:
                template_row_merges.append((merged_range.min_col, merged_range.max_col))
            elif merged_range.min_row > template_row_idx:
                merges_to_shift.append(merged_range)

        for m in merges_to_shift:
            ws.merged_cells.remove(m)

        ws.insert_rows(template_row_idx + 1, amount=num_to_insert)

        for m in merges_to_shift:
            new_min_row = m.min_row + num_to_insert
            new_max_row = m.max_row + num_to_insert
            ws.merge_cells(start_row=new_min_row, start_column=m.min_col,
                           end_row=new_max_row, end_column=m.max_col)

        for i in range(num_to_insert):
            target_row = template_row_idx + 1 + i
            copy_row_style(ws, template_row_idx, target_row)
            for min_col, max_col in template_row_merges:
                ws.merge_cells(start_row=target_row, start_column=min_col, 
                               end_row=target_row, end_column=max_col)

    for i, row_data in enumerate(data_rows):
        current_row = template_row_idx + i
        for key, val in row_data.items():
            key_lower = key.lower()
            if key_lower == 'lab' and (val is None or val == ""):
                val = 0

            if key_lower in col_map:
                col_idx = col_map[key_lower]
                safe_write(ws, current_row, col_idx, val)

    last_data_row = template_row_idx + len(data_rows) - 1
    total_row_idx = None

    for r_idx in range(last_data_row + 1, last_data_row + 10):
        val = ws.cell(row=r_idx, column=2).value 
        if val and "TOTAL" in str(val).upper():
            total_row_idx = r_idx
            break
    
    if not total_row_idx:
        total_row_idx = last_data_row + 1

    return total_row_idx, col_map


@login_required
def previewWorkload(request):
    target_instructor_id = request.GET.get('instructor')
    
    if target_instructor_id:
        instructor = get_object_or_404(Instructor, instructorId=target_instructor_id)
    else:
        user = request.user
        instructor = None
        login_entry = UserLogin.objects.filter(user=user).select_related("instructor").first()
        if login_entry:
            instructor = login_entry.instructor
    
    if not instructor:
        return redirect('instructorLoad')

    semester_id = request.GET.get("semester")
    if semester_id:
        current_semester = Semester.objects.filter(pk=semester_id).first()
    else:
        current_semester = Semester.objects.order_by('-academicYear', 'term').first()

    curriculum = Curriculum.objects.filter(isActive=True).order_by('-createdAt').first()
    
    candidates = Instructor.objects.filter(
        userlogin__user__roles__name='deptHead',
        userlogin__user__isActive=True
    )

    dept_head_obj = candidates.order_by('userlogin__user__is_superuser').first()
    dept_head_name = dept_head_obj.full_name.upper() if dept_head_obj else "TBA"

    inv_context = get_instructor_involvement(instructor)
    
    admin_h = float(inv_context['inv_admin'] or 0)
    research_h = float(inv_context['inv_research'] or 0)
    extension_h = float(inv_context['inv_extension'] or 0)
    instructional_h = float(inv_context['inv_consultation'] or 0)
    
    all_schedules = Schedule.objects.filter(
        instructor=instructor,
        semester=current_semester,
        status='finalized'
    ).select_related('subject', 'section', 'room')

    reg_schedules_list = []
    over_schedules_list = []
    
    unique_sections = set()
    unique_subjects = set()

    for s in all_schedules:
        unique_sections.add(s.section.sectionId)
        unique_subjects.add(s.subject.subjectId)
        
        if s.isOvertime:
            over_schedules_list.append(s)
        else:
            reg_schedules_list.append(s)
            
    no_of_classes = len(unique_sections)
    no_of_preps = len(unique_subjects)

    def group_and_process(s_list):
        grouped = defaultdict(list)
        for s in s_list:
            key = (s.subject.subjectId, s.section.sectionId) 
            grouped[key].append(s)
        rows = []
        for key, group in grouped.items():
            processed = process_schedule_group(group)
            if processed:
                rows.append(processed)
        return rows

    regular_rows = group_and_process(reg_schedules_list)
    overload_rows = group_and_process(over_schedules_list)

    def sum_prop(rows, prop):
        total = 0
        for r in rows:
            val = r.get(prop, "")
            if val != "":
                try: total += float(val)
                except ValueError: pass
        return total

    reg_lec = sum_prop(regular_rows, 'lec')
    reg_lab = sum_prop(regular_rows, 'lab')
    
    over_lec = sum_prop(overload_rows, 'lec')
    over_lab = sum_prop(overload_rows, 'lab')

    total_involvement = admin_h + research_h + extension_h + instructional_h
    
    involvement_data = {
        'admin': format_number(admin_h),
        'research': format_number(research_h),
        'extension': format_number(extension_h),
        'consultation': format_number(instructional_h),
        'total': format_number(total_involvement),
        'note_text': "COE GAD Coordinator" if float(admin_h) > 0 else "", 
        'no_of_classes': no_of_classes,
        'no_of_preps': no_of_preps
    }

    context = {
        'instructor': instructor,
        'semester': current_semester,
        'semester_text': get_semester_text(current_semester.term) if current_semester else "",
        'curriculum': curriculum,
        'dept_head_name': dept_head_name,
        
        **inv_context, 
        
        'involvement_data': involvement_data,
        
        'regular_rows': regular_rows,
        'regular_totals': {
            'units': format_number(sum_prop(regular_rows, 'units')),
            'lec': format_number(reg_lec),
            'lab': format_number(reg_lab),
        },
        
        'overload_rows': overload_rows, 
        'overload_totals': {
            'units': format_number(sum_prop(overload_rows, 'units')),
            'lec': format_number(over_lec),
            'lab': format_number(over_lab),
        }
    }

    return render(request, 'scheduler/workload_preview.html', context)


@login_required
def exportWorkloadExcel(request):
    if request.method == "POST":
        template_path = os.path.join(settings.BASE_DIR, 'static', 'excel_templates', 'workload_template.xlsx')
        try:
            wb = load_workbook(template_path)
            ws = wb.active 
        except FileNotFoundError:
            return HttpResponse("Template file not found.", status=404)

        def to_number(val):
            if not val or str(val).strip() == "": return None 
            try:
                f = float(val)
                return int(f) if f.is_integer() else f
            except (ValueError, TypeError):
                return val 

        def extract_rows(prefix):
            codes = request.POST.getlist(f'{prefix}_code[]')
            titles = request.POST.getlist(f'{prefix}_title[]')
            units = request.POST.getlist(f'{prefix}_units[]')
            lecs = request.POST.getlist(f'{prefix}_lec[]')
            labs = request.POST.getlist(f'{prefix}_lab[]')
            students = request.POST.getlist(f'{prefix}_students[]')
            times = request.POST.getlist(f'{prefix}_time[]')
            days = request.POST.getlist(f'{prefix}_days[]')
            rooms = request.POST.getlist(f'{prefix}_room[]')
            sections = request.POST.getlist(f'{prefix}_section[]')
            
            rows = []
            for i in range(len(codes)):
                if codes[i].strip() or titles[i].strip(): 
                    rows.append({
                        'code': codes[i], 
                        'description': titles[i], 
                        'units': to_number(units[i]),    
                        'time': times[i], 
                        'days': days[i], 
                        'lec': to_number(lecs[i]),       
                        'lab': to_number(labs[i]),       
                        'students': to_number(students[i]), 
                        'room': rooms[i], 
                        'section': sections[i]
                    })
            return rows

        regular_rows = extract_rows('reg')
        overload_rows = extract_rows('over')

        faculty_name_raw = request.POST.get('header_faculty', 'INSTRUCTOR')
        
        variables = {
            'header_faculty': faculty_name_raw,
            'header_semester': request.POST.get('header_semester'),
            'header_rank': request.POST.get('header_rank'),
            'header_sy': request.POST.get('header_sy'),
            'header_college': request.POST.get('header_college'),
            'header_date': request.POST.get('header_date'),
            
            'inv_admin': to_number(request.POST.get('inv_admin', '')),
            'inv_research': to_number(request.POST.get('inv_research', '')),
            'inv_extension': to_number(request.POST.get('inv_extension', '')),
            'inv_consultation': to_number(request.POST.get('inv_consultation', '')),
            'inv_others': to_number(request.POST.get('inv_others', '')), 
            'inv_total': to_number(request.POST.get('inv_total', '')),
            'inv_note': request.POST.get('inv_note', ''),

            'inv_classes': to_number(request.POST.get('inv_classes', '')),
            'inv_preps': to_number(request.POST.get('inv_preps', '')),

            'sig_faculty': request.POST.get('sig_faculty'),
            'sig_dept_head': request.POST.get('sig_dept_head'),
            'sig_dean': request.POST.get('sig_dean'),
            'sig_vp': request.POST.get('sig_vp'),
            'sig_president': request.POST.get('sig_president'),
        }

        for row in ws.iter_rows():
            for cell in row:
                if cell.value in variables:
                    safe_write(ws, cell.row, cell.column, variables[cell.value])

        reg_total_row_idx, reg_col_map = fill_list_section(ws, regular_rows, start_search_row=1)
        if regular_rows and reg_col_map:
            write_section_totals(ws, reg_total_row_idx, regular_rows, reg_col_map)

        over_total_row_idx, over_col_map = fill_list_section(ws, overload_rows, start_search_row=reg_total_row_idx + 2)
        if overload_rows and over_col_map:
            write_section_totals(ws, over_total_row_idx, overload_rows, over_col_map)

        safe_name = faculty_name_raw.strip().replace(" ", "-").upper()
        filename = f"ACTUAL-TEACHING-LOAD-{safe_name}.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    return HttpResponse("Method not allowed")


@login_required
@has_role('deptHead')
def sectionBlockScheduler(request):
    semesters = Semester.objects.all().order_by('-academicYear', '-term')
    active_semester = Semester.objects.filter(isActive=True).first()
    
    selected_semester_id = request.GET.get('semester')
    current_semester = None

    if selected_semester_id:
        try:
            current_semester = Semester.objects.get(semesterId=selected_semester_id)
        except Semester.DoesNotExist:
            current_semester = active_semester
    else:
        current_semester = active_semester
    finalized = Schedule.objects.filter(
        status='finalized',
        semester=current_semester
    ).select_related('section', 'section__subject')
    
    unique_blocks = set()

    for sched in finalized:
        raw_year = str(sched.section.subject.yearLevel)
        raw_code = str(sched.section.sectionCode).strip()

        if '-' in raw_code:
            block_letter = raw_code.split('-')[-1].strip()
        else:
            block_letter = raw_code

        unique_blocks.add((raw_year, block_letter))

    blocks = []
    for year, letter in sorted(list(unique_blocks)):
        blocks.append({
            'label': f"{year}-{letter}", 
            'value': f"{year}__{letter}"
        })

    selected_val = request.GET.get('block')
    
    valid_values = [b['value'] for b in blocks]
    if (not selected_val or selected_val not in valid_values) and blocks:
        selected_val = blocks[0]['value']
    elif not blocks:
        selected_val = None

    schedules = []
    gen_eds = []
    selected_label = ""

    if selected_val:
        try:
            year, letter = selected_val.split('__', 1)
            selected_label = f"{year}-{letter}"

            schedules = Schedule.objects.filter(
                status='finalized',
                section__subject__yearLevel=year,
                semester=current_semester
            ).filter(
                Q(section__sectionCode=letter) |                 
                Q(section__sectionCode__endswith=f"-{letter}")   
            ).select_related('subject', 'instructor', 'room', 'section')

            gen_eds = GenEdSchedule.objects.filter(
                status='active',
                yearLevel=year,
                sectionCode=letter,
                semester=current_semester
            )

        except ValueError:
            pass

    time_slots = [time(h, 0) for h in range(7, 21)]

    context = {
        'semesters': semesters,
        'selected_semester': current_semester,
        'blocks': blocks,
        'selectedBlock': selected_val,
        'selectedBlockLabel': selected_label,
        'schedules': schedules,
        'gen_eds': gen_eds,
        'days': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'],
        'times': time_slots
    }
    
    return render(request, 'scheduler/sectionBlockScheduler.html', context)


@login_required
@has_role('deptHead')
def roomScheduler(request):
    rooms = Room.objects.filter(isActive=True).order_by('building', 'roomCode')
    
    semesters = Semester.objects.all().order_by('-academicYear', '-term')
    active_semester = Semester.objects.filter(isActive=True).first()
    
    selected_semester_id = request.GET.get('semester')
    current_semester = None

    if selected_semester_id:
        try:
            current_semester = Semester.objects.get(semesterId=selected_semester_id)
        except Semester.DoesNotExist:
            current_semester = active_semester
    else:
        current_semester = active_semester

    selected_room_type = request.GET.get('room_type', '') # 'lab', 'lecture', or ''

    selected_room_id = request.GET.get('room')
    selected_room_label = ""
    schedules = []

    if selected_room_id:
        try:
            current_room = rooms.get(roomId=selected_room_id)
            selected_room_label = f"{current_room.roomCode} - {current_room.building}"
            
            if current_room.type == 'laboratory':
                selected_room_label += " (Lab)"

            schedules_qs = Schedule.objects.filter(
                status='finalized',
                room=current_room,
                semester=current_semester
            ).select_related('subject', 'instructor', 'section')

            for sched in schedules_qs:
                raw = str(sched.section.sectionCode).strip()
                letter = raw.split('-')[-1].strip() if '-' in raw else raw
                sched.formatted_section = f"{sched.subject.yearLevel}{letter}" 
                
                if sched.scheduleType == 'lab':
                    sched.type = "Laboratory"
                else:
                    sched.type = "Lecture"

                schedules.append(sched)

        except Room.DoesNotExist:
            pass

    tba_schedules = []
    
    if current_semester:
        tba_qs = Schedule.objects.filter(
            status='finalized',
            semester=current_semester
        ).filter(
            Q(room__roomCode='TBA') | Q(room__isnull=True)
        ).select_related('subject', 'instructor', 'section').order_by('subject__code')

        if selected_room_type == 'lab':
            tba_qs = tba_qs.filter(scheduleType='lab')
        elif selected_room_type == 'lecture':
            tba_qs = tba_qs.exclude(scheduleType='lab')

        for sched in tba_qs:
            raw = str(sched.section.sectionCode).strip()
            letter = raw.split('-')[-1].strip() if '-' in raw else raw
            sched.formatted_section = f"{sched.subject.yearLevel}{letter}"
            
            if sched.scheduleType == 'lab':
                sched.type = "Laboratory"
            else:
                sched.type = "Lecture"

            tba_schedules.append(sched)

    time_slots = [time(h, 0) for h in range(7, 21)]

    context = {
        'semesters': semesters,
        'selected_semester': current_semester,
        'selected_room_type': selected_room_type, 
        'rooms': rooms,
        'selectedRoomId': int(selected_room_id) if selected_room_id else None,
        'selectedRoomLabel': selected_room_label,
        'schedules': schedules,
        'tba_schedules': tba_schedules,
        'days': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'],
        'times': time_slots
    }
    
    return render(request, 'scheduler/roomScheduler.html', context)


@login_required
@has_role('deptHead')
def getInstructorConflicts(request):
    instructorId = request.GET.get('instructorId')
    currentScheduleId = request.GET.get('scheduleId')
    
    source_view = request.GET.get('source')
    semester_id = request.GET.get('semester') 

    busySlots = []

    try:
        current_sched = Schedule.objects.get(scheduleId=currentScheduleId)
        
        target_semester = current_sched.semester
        if semester_id:
            try:
                target_semester = Semester.objects.get(semesterId=semester_id)
            except Semester.DoesNotExist:
                pass


        def check_instructor_conflicts():
            if instructorId:
                inst_conflicts = Schedule.objects.filter(
                    instructor_id=instructorId,
                    status='finalized',
                    semester=target_semester 
                ).exclude(scheduleId=currentScheduleId).select_related('subject', 'section')

                for c in inst_conflicts:
                    raw = str(c.section.sectionCode).strip()
                    letter = raw.split('-')[-1].strip() if '-' in raw else raw
                    short_sec = f"{c.subject.yearLevel}{letter}"
                    
                    busySlots.append({
                        'day': c.dayOfWeek,
                        'startTime': c.startTime.strftime("%H:%M"),
                        'endTime': c.endTime.strftime("%H:%M"),
                        'reason': f"Instructor is teaching {c.subject.code} - {short_sec}"
                    })

        def check_section_conflicts():
            current_year = current_sched.section.subject.yearLevel
            raw_code = str(current_sched.section.sectionCode).strip()
            block_letter = raw_code.split('-')[-1].strip() if '-' in raw_code else raw_code

            section_conflicts = Schedule.objects.filter(
                status='finalized',
                semester=target_semester,
                section__subject__yearLevel=current_year,
                dayOfWeek__in=['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            ).filter(
                Q(section__sectionCode=block_letter) |
                Q(section__sectionCode__endswith=f"-{block_letter}")
            ).exclude(scheduleId=currentScheduleId).select_related('subject', 'room')

            for s in section_conflicts:
                room_name = s.room.roomCode if s.room else "TBA"
                reason = f"Section {current_year}{block_letter} is taking {s.subject.code} in {room_name}"

                busySlots.append({
                    'day': s.dayOfWeek,
                    'startTime': s.startTime.strftime("%H:%M"),
                    'endTime': s.endTime.strftime("%H:%M"),
                    'reason': reason
                })

        if source_view == 'instructorLoad':
            check_section_conflicts()
            
        elif source_view == 'sectionBlockScheduler':
            check_instructor_conflicts()
            
        elif source_view == 'roomScheduler':
            check_instructor_conflicts()
            check_section_conflicts()
            
        else:
            check_instructor_conflicts()
            check_section_conflicts()

    except Exception as e:
        print(f"Error checking conflicts: {e}")

    return JsonResponse({'busySlots': busySlots})


@login_required
@has_role('deptHead')
def updateScheduleSlot(request):
    try:
        data = json.loads(request.body)
        scheduleId = data.get('scheduleId')
        newDay = data.get('day')
        newStartTime = data.get('startTime')
        newEndTime = data.get('endTime')
        newRoomId = data.get('roomId') 
        newInstructorId = data.get('instructorId')
        forceSwap = data.get('forceSwap', False)

        sched = Schedule.objects.get(scheduleId=scheduleId)
        
        if newInstructorId == 'UNASSIGN':
            target_instructor = None
        elif newInstructorId:
            target_instructor = Instructor.objects.get(instructorId=newInstructorId)
        else:
            target_instructor = sched.instructor 

        if newRoomId == 'TBA':
             try: target_room = Room.objects.get(roomCode='TBA')
             except: target_room = None
        elif newRoomId:
             target_room = Room.objects.get(roomId=newRoomId)
        else:
             is_time_change = (newDay is not None) or (newStartTime is not None)
             
             if is_time_change:
                 try: target_room = Room.objects.get(roomCode='TBA')
                 except: target_room = None
             else:
                 target_room = sched.room

        is_tba = (target_room is None) or (target_room.roomCode == 'TBA')

        if is_tba:
             sched.room = target_room
             if newInstructorId is not None: sched.instructor = target_instructor
             if newDay: sched.dayOfWeek = newDay
             if newEndTime: sched.endTime = newEndTime
             
             if newStartTime: 
                 sched.startTime = newStartTime
                 
                 end_dt = datetime.strptime(newEndTime, "%H:%M").time()
                 is_weekend = newDay in ['Saturday', 'Sunday']
                 is_past_5pm = end_dt > time(17, 0)
                 
                 if is_past_5pm or is_weekend:
                     sched.isOvertime = True
                 else:
                     sched.isOvertime = False

             sched.save()
             return JsonResponse({'success': True})

        if target_instructor:
            instructor_conflict = Schedule.objects.filter(
                instructor=target_instructor,
                dayOfWeek=newDay,
                status='finalized',
                semester=sched.semester 
            ).exclude(scheduleId=scheduleId).filter(
                startTime__lt=newEndTime,
                endTime__gt=newStartTime
            ).select_related('subject', 'section').first()

            if instructor_conflict:
                c_raw = str(instructor_conflict.section.sectionCode).strip()
                c_letter = c_raw.split('-')[-1].strip() if '-' in c_raw else c_raw
                return JsonResponse({
                    'success': False, 
                    'message': f"Instructor Conflict! {target_instructor.full_name} is already teaching {instructor_conflict.subject.code} ({instructor_conflict.subject.yearLevel}{c_letter})."
                })

        current_year = sched.section.subject.yearLevel
        raw_code = str(sched.section.sectionCode).strip()
        block_letter = raw_code.split('-')[-1].strip() if '-' in raw_code else raw_code
        formatted_block = f"{current_year}{block_letter}"

        section_conflict = Schedule.objects.filter(
            status='finalized',
            dayOfWeek=newDay,
            section__subject__yearLevel=current_year,
            semester=sched.semester
        ).filter(
            Q(section__sectionCode=block_letter) |
            Q(section__sectionCode__endswith=f"-{block_letter}")
        ).exclude(scheduleId=scheduleId).filter(
            startTime__lt=newEndTime,
            endTime__gt=newStartTime
        ).select_related('subject', 'room').first()

        if section_conflict:
            conflict_room = section_conflict.room.roomCode if section_conflict.room else "TBA"
            return JsonResponse({
                'success': False,
                'message': f"Section Conflict! Section {formatted_block} is already taking '{section_conflict.subject.code}' in {conflict_room}."
            })

        if target_room: 
            room_conflict = Schedule.objects.filter(
                room=target_room,
                dayOfWeek=newDay,
                status='finalized',
                semester=sched.semester
            ).exclude(scheduleId=scheduleId).filter(
                startTime__lt=newEndTime,
                endTime__gt=newStartTime
            ).select_related('subject', 'section').first()

            if room_conflict:
                if forceSwap:
                    try: tba = Room.objects.get(roomCode="TBA")
                    except: tba = None
                    room_conflict.room = tba
                    room_conflict.save()
                else:
                    c_raw = str(room_conflict.section.sectionCode).strip()
                    c_letter = c_raw.split('-')[-1].strip() if '-' in c_raw else c_raw
                    return JsonResponse({
                        'success': False,
                        'needSwapConfirmation': True,
                        'message': f"Slot occupied by {room_conflict.subject.code} - {room_conflict.subject.yearLevel}{c_letter}. Swap and move them to TBA?"
                    })

        if newInstructorId is not None: sched.instructor = target_instructor
        sched.room = target_room
        sched.dayOfWeek = newDay
        sched.startTime = newStartTime
        sched.endTime = newEndTime

        end_dt = datetime.strptime(newEndTime, "%H:%M").time()
        is_weekend = newDay in ['Saturday', 'Sunday']
        is_past_5pm = end_dt > time(17, 0)
        
        if is_past_5pm or is_weekend:
            sched.isOvertime = True
        else:
            sched.isOvertime = False

        sched.save()
        
        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})
    

@login_required
def getInstructorLoadStats(request):
    instructor_id = request.GET.get('instructorId')
    semester_id = request.GET.get('semesterId')

    if not instructor_id:
        return JsonResponse({'success': False, 'message': 'Missing instructorId'})

    try:
        instructor = Instructor.objects.get(instructorId=instructor_id)
        config = InstructorSchedulingConfiguration.objects.filter(is_active=True).first()
        
        reg_limit = 15.0      
        overload_cap = 0.0    
        emp_type = instructor.employmentType 

        if emp_type == 'permanent':
            if instructor.designation:
                reg_limit = float(instructor.designation.instructionHours)
                overload_cap = config.overload_limit_with_designation if config else 9.0
            else:
                reg_limit = float(instructor.rank.instructionHours) if instructor.rank else 18.0
                overload_cap = config.overload_limit_no_designation if config else 12.0
        elif emp_type == 'part-time':
            reg_limit = config.part_time_normal_limit if config else 15.0
            overload_cap = config.part_time_overload_limit if config else 0.0
        elif emp_type == 'overload':
            reg_limit = config.pure_overload_normal_limit if config else 0.0
            overload_cap = config.pure_overload_max_limit if config else 12.0
        else:
            reg_limit = 0.0
            overload_cap = 0.0

        max_total_limit = reg_limit + overload_cap

        if semester_id:
            try:
                target_semester = Semester.objects.get(semesterId=semester_id)
            except Semester.DoesNotExist:
                target_semester = Semester.objects.filter(isActive=True).first()
        else:
            target_semester = Semester.objects.filter(isActive=True).first()
        
        schedules = Schedule.objects.filter(
            instructor=instructor,
            semester=target_semester, 
            status='finalized' 
        ).select_related('subject', 'section')
        
        current_normal_load = 0.0
        current_overload_load = 0.0
        processed_load_keys = set()
        
        evening_cutoff = time(17, 0)

        for s in schedules:
            if s.scheduleType == 'lab':
                units = float(s.subject.labHours)
            else:
                units = float(s.subject.lectureHours)
                
            load_key = (s.section.sectionId, s.scheduleType)
            
            if load_key not in processed_load_keys:
                is_evening = False
                if s.startTime and s.startTime >= evening_cutoff:
                    is_evening = True
                
                is_weekend = s.dayOfWeek in ['Saturday', 'Sunday']

                if is_evening or is_weekend:
                    current_overload_load += units
                else:
                    current_normal_load += units
                
                processed_load_keys.add(load_key)

        if current_normal_load > reg_limit:
            excess = current_normal_load - reg_limit
            current_overload_load += excess
            current_normal_load = reg_limit
            
        current_load = current_normal_load + current_overload_load
        
        if current_load > max_total_limit:
            load_status_color = "bg-red-500"
        elif current_load > reg_limit:
            load_status_color = "bg-amber-500"
        else:
            load_status_color = "bg-emerald-500"

        return JsonResponse({
            'success': True,
            'stats': {
                'normal_load': round(current_normal_load, 2),
                'overload_load': round(current_overload_load, 2),
                'current_load': round(current_load, 2),
                'reg_limit': round(reg_limit, 2),
                'overload_cap': round(overload_cap, 2),
                'max_limit': round(max_total_limit, 2),
                'load_color': load_status_color
            }
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@has_role('deptHead')
def instructorLoad(request):
    instructors = Instructor.objects.filter(
        userlogin__user__isActive=True
    ).distinct().order_by('instructorId')
    
    def get_instructor_name(inst):
        login = inst.userlogin_set.select_related('user').first()
        if login and login.user:
            return login.user.get_full_name() 
        return inst.instructorId

    instructor_list = []
    for inst in instructors:
        instructor_list.append({
            'instructorId': inst.instructorId,
            'full_name': get_instructor_name(inst)
        })
    instructor_list.sort(key=lambda x: x['full_name'])

    semesters = Semester.objects.all().order_by('-academicYear', '-term')
    active_semester = Semester.objects.filter(isActive=True).first()
    
    selected_semester_id = request.GET.get('semester')
    current_semester = None

    if selected_semester_id:
        try:
            current_semester = Semester.objects.get(semesterId=selected_semester_id)
        except Semester.DoesNotExist:
            current_semester = active_semester
    else:
        current_semester = active_semester

    selected_instructor_id = request.GET.get('instructor')
    selected_instructor_label = ""
    schedules = []
    
    current_load = 0.0
    current_normal_load = 0.0  
    current_overload_load = 0.0
    reg_limit = 15.0      
    overload_cap = 0.0    
    max_total_limit = 0.0 
    load_status_color = "bg-emerald-500"

    match_scores = {}
    top_matches = [] 

    config = InstructorSchedulingConfiguration.objects.filter(is_active=True).first()

    if selected_instructor_id and current_semester:
        try:
            current_instructor = instructors.get(instructorId=selected_instructor_id)
            selected_instructor_label = get_instructor_name(current_instructor)
            
            raw_matches = InstructorSubjectMatch.objects.filter(
                instructor=current_instructor
            ).select_related('subject', 'latestHistory').order_by('-generatedAt')

            seen_subjects = set()
            for m in raw_matches:
                if m.subject.subjectId in seen_subjects:
                    continue
                seen_subjects.add(m.subject.subjectId)

                final_score = 0
                if m.latestHistory:
                    c_score = m.latestHistory.confidenceScore
                    final_score = round(c_score * 100, 1) if c_score <= 1.0 else round(c_score, 1)
                
                match_scores[m.subject.subjectId] = final_score
                if final_score > 0: 
                    top_matches.append({
                        'code': m.subject.code,
                        'score': final_score,
                        'badge_color': 'green' if final_score >= 85 else 'yellow' if final_score >= 60 else 'gray'
                    })
            top_matches.sort(key=lambda x: x['score'], reverse=True)

            emp_type = (current_instructor.employmentType or "").lower().strip()

            if emp_type == 'permanent':
                has_designation = (
                    current_instructor.designation and 
                    current_instructor.designation.name.strip().upper() != 'N/A'
                )

                if has_designation:
                    reg_limit = float(current_instructor.designation.instructionHours)
                    overload_cap = config.overload_limit_with_designation if config else 9.0
                else:
                    has_rank = (
                        current_instructor.rank and 
                        current_instructor.rank.name.strip().upper() != 'N/A'
                    )
                    
                    if has_rank:
                        reg_limit = float(current_instructor.rank.instructionHours)
                    else:
                        reg_limit = 18.0

                    overload_cap = config.overload_limit_no_designation if config else 12.0
            
            elif emp_type == 'part-time':
                reg_limit = config.part_time_normal_limit if config else 15.0
                overload_cap = config.part_time_overload_limit if config else 0.0
            
            elif emp_type == 'overload':
                reg_limit = config.pure_overload_normal_limit if config else 0.0
                overload_cap = config.pure_overload_max_limit if config else 12.0
            
            else:
                reg_limit = 0.0
                overload_cap = 0.0

            max_total_limit = reg_limit + overload_cap

            schedules_qs = Schedule.objects.filter(
                status='finalized',
                instructor=current_instructor,
                semester=current_semester
            ).select_related('subject', 'section', 'room')
            
            evening_cutoff = time(17, 0)

            for sched in schedules_qs:
                raw = str(sched.section.sectionCode).strip()
                letter = raw.split('-')[-1].strip() if '-' in raw else raw
                sched.formatted_section = f"{sched.subject.yearLevel}{letter}" 
                
                if sched.scheduleType == 'lab':
                    sched.type = "Laboratory"
                else:
                    sched.type = "Lecture"
                
                start_min = sched.startTime.hour * 60 + sched.startTime.minute
                end_min = sched.endTime.hour * 60 + sched.endTime.minute
                duration_minutes = end_min - start_min
                
                slot_units = duration_minutes / 60.0
                
                is_weekend = sched.dayOfWeek in ['Saturday', 'Sunday']
                
                cutoff_min = 17 * 60

                if is_weekend or end_min > cutoff_min:
                    current_overload_load += slot_units
                else:
                    current_normal_load += slot_units

                sched.match_score = match_scores.get(sched.subject.subjectId, 0)
                schedules.append(sched)

            if current_normal_load > reg_limit:
                excess = current_normal_load - reg_limit
                current_overload_load += excess
                current_normal_load = reg_limit

            current_load = current_normal_load + current_overload_load

            if current_load > max_total_limit:
                load_status_color = "bg-red-500"
            elif current_load > reg_limit:
                load_status_color = "bg-amber-500"
            else:
                load_status_color = "bg-emerald-500"

        except Instructor.DoesNotExist:
            pass

    unassigned_schedules = []
    if current_semester:
        unassigned_qs = Schedule.objects.filter(
            status='finalized',
            instructor__isnull=True,
            semester=current_semester
        ).select_related('subject', 'section', 'room').order_by('subject__code')

        for sched in unassigned_qs:
            raw = str(sched.section.sectionCode).strip()
            letter = raw.split('-')[-1].strip() if '-' in raw else raw
            sched.formatted_section = f"{sched.subject.yearLevel}{letter}"
            
            if sched.scheduleType == 'lab':
                sched.type = "Laboratory"
            else:
                sched.type = "Lecture"

            if selected_instructor_id:
                sched.match_score = match_scores.get(sched.subject.subjectId, 0)
            else:
                sched.match_score = 0
            unassigned_schedules.append(sched)

    time_slots = [time(h, 0) for h in range(7, 21)]

    context = {
        'semesters': semesters,
        'selected_semester': current_semester,
        'instructors': instructor_list,
        'selectedInstructorId': selected_instructor_id,
        'selectedInstructorLabel': selected_instructor_label,
        'schedules': schedules,
        'unassigned_schedules': unassigned_schedules,
        'top_matches': top_matches,
        'days': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday'],
        'times': time_slots,
        'current_load': round(current_load, 2),
        'normal_load': round(current_normal_load, 2),
        'overload_load': round(current_overload_load, 2),
        'reg_limit': round(reg_limit, 2),
        'overload_cap': round(overload_cap, 2),
        'max_limit': round(max_total_limit, 2),
        'load_color': load_status_color
    }
    
    return render(request, 'scheduler/instructorLoad.html', context)

@login_required
@has_role('deptHead')
def previewRoomSchedule(request, roomId, semesterId):
    room = get_object_or_404(Room, pk=roomId)
    semester = get_object_or_404(Semester, pk=semesterId)
    
    deptHead = Instructor.objects.filter(
        userlogin__user__roles__name='deptHead', 
        userlogin__user__isActive=True
    ).order_by('userlogin__user__is_superuser').first()

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    startHour = 7
    endHour = 21
    
    schedules = Schedule.objects.filter(
        room=room, 
        semester=semester, 
        status='finalized'
    )
    
    scheduleMap = {}
    for sched in schedules:
        key = (sched.dayOfWeek, sched.startTime.strftime("%H:%M"))
        scheduleMap[key] = sched

    gridRows = []
    skipCounts = {day: 0 for day in days}
    
    today = datetime.today()
    currentDt = datetime.combine(today, time(startHour, 0))
    endDt = datetime.combine(today, time(endHour, 0))
    lunchStart = datetime.combine(today, time(12, 0))

    while currentDt < endDt:
        currentTimeStr = currentDt.strftime("%H:%M")
        
        if currentDt.hour == 12:
            for day in days:
                skipCounts[day] = 0
                
            gridRows.append({'isLunch': True, 'label': '12:00-1:00', 'cells': []})
            currentDt += timedelta(hours=1)
            continue

        slotEnd = currentDt + timedelta(minutes=30)
        
        startStr = currentDt.strftime('%I:%M').lstrip('0')
        endStr = slotEnd.strftime('%I:%M').lstrip('0')
        timeLabel = f"{startStr}-{endStr}"
        
        rowObj = {'isLunch': False, 'timeLabel': timeLabel, 'cells': []}

        for day in days:
            if skipCounts[day] > 0:
                rowObj['cells'].append({'type': 'span'})
                skipCounts[day] -= 1
            else:
                schedKey = (day, currentTimeStr)
                if schedKey in scheduleMap:
                    sched = scheduleMap[schedKey]
                    
                    sStart = datetime.combine(today, sched.startTime)
                    sEnd = datetime.combine(today, sched.endTime)
                    
                    if sStart < lunchStart and sEnd > lunchStart:
                        sEnd = lunchStart
                    
                    if sEnd > endDt:
                        sEnd = endDt

                    durationMins = (sEnd - sStart).total_seconds() / 60
                    rowSpan = int(durationMins / 30)
                    
                    rawSection = str(sched.section.sectionCode).strip()
                    if '-' in rawSection:
                        sectionLetter = rawSection.split('-')[-1].strip()
                    else:
                        sectionLetter = rawSection
                    
                    formattedSection = f"{sched.subject.yearLevel}{sectionLetter}"

                    rowObj['cells'].append({
                        'type': 'event',
                        'rowSpan': rowSpan,
                        'subjectCode': sched.subject.code,
                        'section': formattedSection,
                        'schedType': sched.scheduleType.title(),
                        'instructorName': sched.instructor.full_name
                    })
                    skipCounts[day] = rowSpan - 1
                else:
                    rowObj['cells'].append({'type': 'empty'})

        gridRows.append(rowObj)
        currentDt += timedelta(minutes=30)

    context = {
        'room': room,
        'semester': semester,
        'deptHead': deptHead,
        'days': days,
        'gridRows': gridRows,
    }
    return render(request, 'scheduler/previewRoomSchedule.html', context)


@login_required
@has_role('deptHead')
def previewSectionBlockSchedule(request, blockStr, semesterId):
    try:
        year, letter = blockStr.split('__', 1)
        label = f"{year}-{letter}"
    except ValueError:
        return redirect('sectionBlockScheduler') 

    semester = get_object_or_404(Semester, pk=semesterId)
    
    deptHead = Instructor.objects.filter(
        userlogin__user__roles__name='deptHead', 
        userlogin__user__isActive=True
    ).order_by('userlogin__user__is_superuser').first()

    major_schedules = Schedule.objects.filter(
        status='finalized',
        section__subject__yearLevel=year,
        semester=semester
    ).filter(
        Q(section__sectionCode=letter) |                 
        Q(section__sectionCode__endswith=f"-{letter}")   
    ).select_related('subject', 'instructor', 'room', 'section')

    gen_eds = GenEdSchedule.objects.filter(
        status='active',
        yearLevel=year,
        sectionCode=letter,
        semester=semester
    )

    scheduleMap = {}
    
    for sched in major_schedules:
        key = (sched.dayOfWeek, sched.startTime.strftime("%H:%M"))
        
        subject_display = f"{sched.subject.code} - {sched.scheduleType.title()}"

        scheduleMap[key] = {
            'startTime': sched.startTime,
            'endTime': sched.endTime,
            'subjectCode': subject_display, # Line 1
            'roomName': sched.room.roomCode if sched.room else "TBA", # Line 2 (Room)
            'instructorName': sched.instructor.full_name, # Line 3
            'isGenEd': False
        }

    for gen in gen_eds:
        key = (gen.dayOfWeek, gen.startTime.strftime("%H:%M"))
        
        subject_display = f"{gen.code} - Lecture"

        scheduleMap[key] = {
            'startTime': gen.startTime,
            'endTime': gen.endTime,
            'subjectCode': subject_display, 
            'roomName': gen.room or "TBA",
            'instructorName': gen.instructorName or "TBA",
            'isGenEd': True
        }

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    gridRows = []
    skipCounts = {day: 0 for day in days}
    
    startHour = 7
    endHour = 21
    
    today = datetime.today()
    currentDt = datetime.combine(today, time(startHour, 0))
    endDt = datetime.combine(today, time(endHour, 0))
    lunchStart = datetime.combine(today, time(12, 0))

    while currentDt < endDt:
        currentTimeStr = currentDt.strftime("%H:%M")
        
        if currentDt.hour == 12:
            for day in days: skipCounts[day] = 0
            gridRows.append({'isLunch': True, 'label': '12:00-1:00', 'cells': []})
            currentDt += timedelta(hours=1)
            continue

        slotEnd = currentDt + timedelta(minutes=30)
        startStr = currentDt.strftime('%I:%M').lstrip('0')
        endStr = slotEnd.strftime('%I:%M').lstrip('0')
        timeLabel = f"{startStr}-{endStr}"
        
        rowObj = {'isLunch': False, 'timeLabel': timeLabel, 'cells': []}

        for day in days:
            if skipCounts[day] > 0:
                rowObj['cells'].append({'type': 'span'})
                skipCounts[day] -= 1
            else:
                schedKey = (day, currentTimeStr)
                if schedKey in scheduleMap:
                    data = scheduleMap[schedKey]
                    
                    sStart = datetime.combine(today, data['startTime'])
                    sEnd = datetime.combine(today, data['endTime'])
                    
                    if sStart < lunchStart and sEnd > lunchStart: sEnd = lunchStart
                    if sEnd > endDt: sEnd = endDt

                    durationMins = (sEnd - sStart).total_seconds() / 60
                    rowSpan = int(durationMins / 30)
                    
                    rowObj['cells'].append({
                        'type': 'event',
                        'rowSpan': rowSpan,
                        'line1': data['subjectCode'],  
                        'line2': data['roomName'],
                        'line3': data['instructorName'],
                        'isGenEd': data['isGenEd']
                    })
                    skipCounts[day] = rowSpan - 1
                else:
                    rowObj['cells'].append({'type': 'empty'})

        gridRows.append(rowObj)
        currentDt += timedelta(minutes=30)

    context = {
        'blockLabel': label,
        'semester': semester,
        'deptHead': deptHead,
        'days': days,
        'gridRows': gridRows,
    }
    return render(request, 'scheduler/previewSectionBlockSchedule.html', context)